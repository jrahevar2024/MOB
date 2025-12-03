"""
Document processing utilities for extracting text from various file formats
"""
import logging
import os
import subprocess
import tempfile
from io import BytesIO
from typing import Optional, Dict

logger = logging.getLogger(__name__)

# Check for optional dependencies
try:
    import PyPDF2
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False
    logger.warning("PyPDF2 not available. PDF processing will be disabled.")

try:
    from docx import Document as DocxDocument
    DOCX_AVAILABLE = True
except ImportError:
    DOCX_AVAILABLE = False
    logger.warning("python-docx not available. DOCX processing will be disabled.")

try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logger.warning("openpyxl not available. Excel processing will be disabled.")


def extract_text_from_pdf(file_bytes: bytes) -> str:
    """Extract text from PDF file"""
    if not PDF_AVAILABLE:
        return "PDF processing not available. Please install PyPDF2."
    try:
        pdf_file = BytesIO(file_bytes)
        pdf_reader = PyPDF2.PdfReader(pdf_file)
        text = ""
        for page in pdf_reader.pages:
            text += page.extract_text() + "\n"
        return text
    except Exception as e:
        logger.error(f"Error extracting text from PDF: {str(e)}")
        return f"Error extracting text from PDF: {str(e)}"


def extract_text_from_docx(file_bytes: bytes) -> str:
    """Extract text from DOCX file"""
    if not DOCX_AVAILABLE:
        return "DOCX processing not available. Please install python-docx."
    try:
        docx_file = BytesIO(file_bytes)
        doc = DocxDocument(docx_file)
        text = "\n".join([paragraph.text for paragraph in doc.paragraphs])
        return text
    except Exception as e:
        logger.error(f"Error extracting text from DOCX: {str(e)}")
        return f"Error extracting text from DOCX: {str(e)}"


def extract_text_from_doc(file_bytes: bytes) -> str:
    """Extract text from DOC file (old Word format) using LibreOffice conversion"""
    tmp_doc_path = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix='.doc') as tmp_doc:
            tmp_doc.write(file_bytes)
            tmp_doc_path = tmp_doc.name
        
        output_dir = os.path.dirname(tmp_doc_path)
        
        result = subprocess.run(
            ['libreoffice', '--headless', '--convert-to', 'txt:Text', 
             '--outdir', output_dir, tmp_doc_path],
            capture_output=True,
            text=True,
            timeout=30
        )
        
        if result.returncode == 0:
            txt_file_path = tmp_doc_path.replace('.doc', '.txt')
            if os.path.exists(txt_file_path):
                with open(txt_file_path, 'r', encoding='utf-8', errors='ignore') as f:
                    text = f.read()
                try:
                    os.unlink(txt_file_path)
                except:
                    pass
                return text
            else:
                raise Exception("LibreOffice conversion succeeded but output file not found")
        else:
            error_msg = result.stderr or result.stdout or "Unknown error"
            raise Exception(f"LibreOffice conversion failed: {error_msg}")
            
    except subprocess.TimeoutExpired:
        return (
            "Error: Document conversion timed out.\n\n"
            "Please try converting your .doc file to .docx format manually and upload it again."
        )
    except FileNotFoundError:
        return (
            "DOC file processing requires LibreOffice to be installed.\n\n"
            "Install LibreOffice:\n"
            "- macOS: brew install --cask libreoffice\n"
            "- Linux: sudo apt-get install libreoffice\n"
            "- Windows: Download from https://www.libreoffice.org/\n\n"
            "After installing LibreOffice, restart the application.\n\n"
            "Alternatively, you can convert your .doc file to .docx format and upload it again."
        )
    except Exception as e:
        logger.error(f"Error extracting text from DOC: {str(e)}")
        return (
            f"Error processing .doc file: {str(e)}\n\n"
            "Please try one of the following:\n"
            "1. Convert your .doc file to .docx format and upload it again\n"
            "2. Install LibreOffice for automatic conversion\n"
            "3. Use an online converter to convert .doc to .docx"
        )
    finally:
        if tmp_doc_path and os.path.exists(tmp_doc_path):
            try:
                os.unlink(tmp_doc_path)
            except:
                pass


def extract_text_from_excel(file_bytes: bytes) -> str:
    """Extract text from Excel file"""
    if not EXCEL_AVAILABLE:
        return "Excel processing not available. Please install openpyxl."
    try:
        excel_file = BytesIO(file_bytes)
        workbook = openpyxl.load_workbook(excel_file)
        text = ""
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text += f"\n--- Sheet: {sheet_name} ---\n"
            for row in sheet.iter_rows(values_only=True):
                row_text = "\t".join([str(cell) if cell is not None else "" for cell in row])
                if row_text.strip():
                    text += row_text + "\n"
        return text
    except Exception as e:
        logger.error(f"Error extracting text from Excel: {str(e)}")
        return f"Error extracting text from Excel: {str(e)}"


def detect_file_type_by_content(file_bytes: bytes) -> str:
    """Detect file type by checking magic bytes/file signature"""
    if len(file_bytes) < 4:
        return "unknown"
    
    # Check for ZIP-based formats (DOCX, XLSX are ZIP files)
    if file_bytes[:2] == b'PK':
        try:
            import zipfile
            with zipfile.ZipFile(BytesIO(file_bytes)) as zf:
                if 'word/document.xml' in zf.namelist():
                    return "docx"
                elif any(f.startswith('xl/') for f in zf.namelist()):
                    return "xlsx"
        except:
            pass
    
    # Check for PDF
    if file_bytes[:4] == b'%PDF':
        return "pdf"
    
    # Check for old Word format (OLE compound document)
    if file_bytes[:8] == b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1':
        return "doc"
    
    # Check if it's plain text (readable ASCII/UTF-8)
    try:
        file_bytes[:1000].decode('utf-8')
        return "text"
    except:
        pass
    
    return "unknown"


def extract_text_from_file(file_bytes: bytes, file_type: str, file_name: str = "") -> str:
    """Extract text from various file types"""
    file_type_lower = file_type.lower()
    file_name_lower = file_name.lower() if file_name else ""
    
    # First, try to detect actual file type by content (magic bytes)
    detected_type = detect_file_type_by_content(file_bytes)
    
    # Check file extension first (more reliable than MIME type)
    if file_name_lower.endswith(".pdf") or file_type_lower.endswith("pdf") or detected_type == "pdf":
        return extract_text_from_pdf(file_bytes)
    elif file_name_lower.endswith(".doc") or detected_type == "doc":
        return extract_text_from_doc(file_bytes)
    elif file_name_lower.endswith(".docx") or file_type_lower == "application/vnd.openxmlformats-officedocument.wordprocessingml.document" or detected_type == "docx":
        try:
            import zipfile
            zipfile.ZipFile(BytesIO(file_bytes))
            return extract_text_from_docx(file_bytes)
        except Exception as e:
            logger.warning(f"File {file_name} identified as DOCX but is not a valid zip file. Detected type: {detected_type}")
            
            if detected_type == "text":
                try:
                    return file_bytes.decode('utf-8')
                except:
                    pass
            
            logger.info(f"Trying to process {file_name} as DOC file")
            return extract_text_from_doc(file_bytes)
    elif file_type_lower == "application/msword":
        return extract_text_from_doc(file_bytes)
    elif file_name_lower.endswith(("xlsx", "xls")) or file_type_lower in ["application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             "application/vnd.ms-excel"] or detected_type == "xlsx":
        return extract_text_from_excel(file_bytes)
    elif file_type_lower.startswith("text/") or detected_type == "text" or file_name_lower.endswith(".txt"):
        try:
            return file_bytes.decode('utf-8')
        except UnicodeDecodeError:
            try:
                return file_bytes.decode('latin-1')
            except Exception as e:
                return f"Error decoding text file: {str(e)}"
    else:
        if detected_type == "text":
            try:
                return file_bytes.decode('utf-8')
            except:
                pass
        return f"Unsupported file type: {file_type}. Detected type: {detected_type}. Supported types: PDF, DOC, DOCX, TXT, XLSX"


def process_document(file_bytes: bytes, file_name: str, file_type: str, max_size: int = 50 * 1024 * 1024, max_text_length: int = 100 * 1024) -> Optional[Dict]:
    """
    Process an uploaded file and return document info with memory limits
    
    Args:
        file_bytes: The file content as bytes
        file_name: The file name
        file_type: The MIME type or file extension
        max_size: Maximum file size in bytes (default: 50MB)
        max_text_length: Maximum extracted text length in bytes (default: 100KB)
    
    Returns:
        Dictionary with document info or None if processing failed
    """
    try:
        # Limit file size to prevent memory issues
        if len(file_bytes) > max_size:
            error_msg = f"File {file_name} is too large ({len(file_bytes) / 1024 / 1024:.1f}MB). Maximum size is {max_size / 1024 / 1024}MB."
            logger.error(error_msg)
            return {
                "status": "error",
                "error": error_msg
            }
        
        # Extract text from file
        text_content = extract_text_from_file(file_bytes, file_type, file_name)
        
        # Limit extracted text content to prevent memory issues
        if len(text_content) > max_text_length:
            text_content = text_content[:max_text_length] + f"\n\n[Document content truncated - extracted {max_text_length} characters of {len(text_content)} total]"
            logger.warning(f"Document {file_name} content truncated from {len(text_content)} to {max_text_length} characters")
        
        return {
            "status": "success",
            "name": file_name,
            "type": file_type,
            "size": len(file_bytes),
            "content": text_content,
            "preview": text_content[:500] + "..." if len(text_content) > 500 else text_content
        }
    except MemoryError:
        error_msg = f"Not enough memory to process file {file_name}. Please try a smaller file."
        logger.error(error_msg)
        return {
            "status": "error",
            "error": error_msg
        }
    except Exception as e:
        logger.error(f"Error processing uploaded file: {str(e)}")
        return {
            "status": "error",
            "error": f"Error processing file: {str(e)}"
        }

