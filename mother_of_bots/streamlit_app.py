import streamlit as st
import asyncio
import time
import logging
import os
import uuid
import nest_asyncio
import httpx
import gc
from dotenv import load_dotenv
from io import BytesIO
from typing import Optional, Dict, List

# Apply nest_asyncio to allow nested event loops
nest_asyncio.apply()

# Load environment variables
load_dotenv()

# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Document processing imports
try:
    import PyPDF2
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False
    logger.warning("PyPDF2 not available. PDF processing will be disabled.")

try:
    from docx import Document as DocxDocument
    DOCX_AVAILABLE = True
except ImportError:
    DOCX_AVAILABLE = False
    logger.warning("python-docx not available. DOCX processing will be disabled.")

try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logger.warning("openpyxl not available. Excel processing will be disabled.")

# Note: .doc files are processed using LibreOffice (system dependency)
# pypandoc is not used as it doesn't support .doc format directly

# Configuration from environment variables
GCP_PROJECT_ID = os.getenv("GCP_PROJECT_ID", "motherofbots")
GCP_LOCATION = os.getenv("GCP_LOCATION", "us-central1")
GEMINI_MODEL = os.getenv("GEMINI_MODEL", "gemini-2.5-flash")

# Flask API configuration
API_BASE_URL = os.getenv("API_BASE_URL", "http://localhost:8000")

# Setup page config
st.set_page_config(
    page_title="Mother of Bots - Multi-Agent Chat Interface",
    page_icon="🤖",
    layout="centered",
    initial_sidebar_state="expanded",
)

# Add custom CSS for better look and feel
st.markdown("""
<style>
.chat-message {
    padding: 1.5rem; 
    border-radius: 0.5rem; 
    margin-bottom: 1rem; 
    display: flex;
    flex-direction: column;
}
.chat-message.user {
    background-color: #1E88E5; /* Bright Blue */
    color: #FFFFFF;
}
.chat-message.assistant {
    background-color: #43A047; /* Vibrant Green */
    color: #FFFFFF;
}
.chat-message.system {
    background-color: #F4511E; /* Deep Orange */
    color: #FFFFFF;
    font-size: 0.85em;
    opacity: 0.95;
}
.chat-message .avatar {
    width: 20%;
}
.chat-message .avatar img {
    max-width: 78px;
    max-height: 78px;
    border-radius: 50%;
    object-fit: cover;
    border: 3px solid #FFFFFF;
}
.chat-message .message {
    width: 100%;
    padding: 0 1.5rem;
}
h1 {
    color: #FFD600; /* Vivid Yellow */
    text-shadow: 1px 1px 3px rgba(0,0,0,0.5);
}
/* Requirements analysis styling */
.requirements-analysis h3 {
    color: #00E676; /* Neon Green */
    margin-top: 0.8rem;
    margin-bottom: 0.3rem;
    font-size: 1.1rem;
}
.requirements-analysis ul {
    margin-top: 0.2rem;
}
/* Code generation styling */
.code-generation-output {
    margin-top: 1rem;
    border-left: 4px solid #00E5FF; /* Electric Blue */
    padding-left: 1rem;
}
.code-generation-output h2 {
    color: #00E5FF;
    font-size: 1.2rem;
    margin-top: 1rem;
    margin-bottom: 0.5rem;
}
/* Custom styling for different code types */
.code-generation-output h2:contains("Backend") {
    color: #FF3D00; /* Fiery Red */
}
.code-generation-output h2:contains("UI") {
    color: #FFD600; /* Vivid Yellow */
}
.code-generation-output pre {
    background-color: #263238; /* Charcoal */
    color: #FFFFFF;
    padding: 1rem;
    border-radius: 8px;
    overflow-x: auto;
}
/* Different syntax highlighting styles based on code type */
.language-python {
    border-left: 4px solid #4CAF50; /* Fresh Green */
}
.language-jsx, .language-javascript, .language-tsx {
    border-left: 4px solid #FFAB00; /* Amber */
}
.chat-input-container {
    display: flex;
    align-items: center;
    background-color: #212121;
    padding: 1rem;
    border-radius: 0.5rem;
}
.chat-input {
    flex: 1;
    background: #424242;
    color: #FFFFFF;
    border: none;
    padding: 0.75rem 1rem;
    border-radius: 0.3rem;
}
.action-buttons {
    display: flex;
    gap: 0.5rem;
    margin-top: 0.5rem;
}
.action-buttons button {
    background: #00E5FF;
    color: #000000;
    border: none;
    padding: 0.5rem 1rem;
    border-radius: 0.3rem;
    font-weight: bold;
    cursor: pointer;
    transition: background 0.3s ease;
}
.action-buttons button:hover {
    background: #00B8D4;
}
</style>

""", unsafe_allow_html=True)

# Initialize session state variables
if 'agent' not in st.session_state:
    st.session_state.agent = None
    st.session_state.agent_running = False
    st.session_state.messages = []
    st.session_state.user_id = f"user_{uuid.uuid4()}"
    st.session_state.waiting_for_response = False
    
    # Using Flask API mode (no agent_type needed)
    st.session_state.agent_type = "flask"
    
    st.session_state.show_analysis = True  # Show requirements analysis by default
    st.session_state.auto_generate_code = True  # Automatically generate code after analysis
    st.session_state.deploy_services = True  # Enable automatic deployment by default
    st.session_state.deployer_agent = None  # Store deployer agent for stopping services
    st.session_state.backend_url = None  # Backend URL for deployed services
    st.session_state.frontend_url = None  # Frontend URL for deployed services
    st.session_state.uploaded_documents = []  # Store uploaded documents

# Document processing functions
def extract_text_from_pdf(file_bytes: bytes) -> str:
    """Extract text from PDF file"""
    if not PDF_AVAILABLE:
        return "PDF processing not available. Please install PyPDF2."
    try:
        pdf_file = BytesIO(file_bytes)
        pdf_reader = PyPDF2.PdfReader(pdf_file)
        text = ""
        for page in pdf_reader.pages:
            text += page.extract_text() + "\n"
        return text
    except Exception as e:
        logger.error(f"Error extracting text from PDF: {str(e)}")
        return f"Error extracting text from PDF: {str(e)}"

def extract_text_from_docx(file_bytes: bytes) -> str:
    """Extract text from DOCX file"""
    if not DOCX_AVAILABLE:
        return "DOCX processing not available. Please install python-docx."
    try:
        docx_file = BytesIO(file_bytes)
        doc = DocxDocument(docx_file)
        text = "\n".join([paragraph.text for paragraph in doc.paragraphs])
        return text
    except Exception as e:
        logger.error(f"Error extracting text from DOCX: {str(e)}")
        return f"Error extracting text from DOCX: {str(e)}"

def extract_text_from_doc(file_bytes: bytes) -> str:
    """Extract text from DOC file (old Word format) using LibreOffice conversion"""
    import tempfile
    import subprocess
    
    # Save to temporary file
    tmp_doc_path = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix='.doc') as tmp_doc:
            tmp_doc.write(file_bytes)
            tmp_doc_path = tmp_doc.name
        
        # Get output directory
        output_dir = os.path.dirname(tmp_doc_path)
        
        # Use LibreOffice to convert DOC to TXT
        result = subprocess.run(
            ['libreoffice', '--headless', '--convert-to', 'txt:Text', 
             '--outdir', output_dir, tmp_doc_path],
            capture_output=True,
            text=True,
            timeout=30
        )
        
        if result.returncode == 0:
            # Read the converted text file
            txt_file_path = tmp_doc_path.replace('.doc', '.txt')
            if os.path.exists(txt_file_path):
                with open(txt_file_path, 'r', encoding='utf-8', errors='ignore') as f:
                    text = f.read()
                # Clean up converted file
                try:
                    os.unlink(txt_file_path)
                except:
                    pass
                return text
            else:
                raise Exception("LibreOffice conversion succeeded but output file not found")
        else:
            error_msg = result.stderr or result.stdout or "Unknown error"
            raise Exception(f"LibreOffice conversion failed: {error_msg}")
            
    except subprocess.TimeoutExpired:
        return (
            "Error: Document conversion timed out.\n\n"
            "Please try converting your .doc file to .docx format manually and upload it again."
        )
    except FileNotFoundError:
        return (
            "DOC file processing requires LibreOffice to be installed.\n\n"
            "Install LibreOffice:\n"
            "- macOS: brew install --cask libreoffice\n"
            "- Linux: sudo apt-get install libreoffice\n"
            "- Windows: Download from https://www.libreoffice.org/\n\n"
            "After installing LibreOffice, restart the application.\n\n"
            "Alternatively, you can convert your .doc file to .docx format and upload it again."
        )
    except Exception as e:
        logger.error(f"Error extracting text from DOC: {str(e)}")
        return (
            f"Error processing .doc file: {str(e)}\n\n"
            "Please try one of the following:\n"
            "1. Convert your .doc file to .docx format and upload it again\n"
            "2. Install LibreOffice for automatic conversion\n"
            "3. Use an online converter to convert .doc to .docx"
        )
    finally:
        # Clean up temp doc file
        if tmp_doc_path and os.path.exists(tmp_doc_path):
            try:
                os.unlink(tmp_doc_path)
            except:
                pass

def extract_text_from_excel(file_bytes: bytes) -> str:
    """Extract text from Excel file"""
    if not EXCEL_AVAILABLE:
        return "Excel processing not available. Please install openpyxl."
    try:
        excel_file = BytesIO(file_bytes)
        workbook = openpyxl.load_workbook(excel_file)
        text = ""
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text += f"\n--- Sheet: {sheet_name} ---\n"
            for row in sheet.iter_rows(values_only=True):
                row_text = "\t".join([str(cell) if cell is not None else "" for cell in row])
                if row_text.strip():
                    text += row_text + "\n"
        return text
    except Exception as e:
        logger.error(f"Error extracting text from Excel: {str(e)}")
        return f"Error extracting text from Excel: {str(e)}"

def detect_file_type_by_content(file_bytes: bytes) -> str:
    """Detect file type by checking magic bytes/file signature"""
    if len(file_bytes) < 4:
        return "unknown"
    
    # Check for ZIP-based formats (DOCX, XLSX are ZIP files)
    if file_bytes[:2] == b'PK':
        # Check if it's a DOCX (has word/document.xml)
        try:
            import zipfile
            with zipfile.ZipFile(BytesIO(file_bytes)) as zf:
                if 'word/document.xml' in zf.namelist():
                    return "docx"
                elif any(f.startswith('xl/') for f in zf.namelist()):
                    return "xlsx"
        except:
            pass
    
    # Check for PDF
    if file_bytes[:4] == b'%PDF':
        return "pdf"
    
    # Check for old Word format (OLE compound document)
    if file_bytes[:8] == b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1':
        return "doc"
    
    # Check if it's plain text (readable ASCII/UTF-8)
    try:
        file_bytes[:1000].decode('utf-8')
        return "text"
    except:
        pass
    
    return "unknown"

def extract_text_from_file(file_bytes: bytes, file_type: str, file_name: str = "") -> str:
    """Extract text from various file types"""
    file_type_lower = file_type.lower()
    file_name_lower = file_name.lower() if file_name else ""
    
    # First, try to detect actual file type by content (magic bytes)
    detected_type = detect_file_type_by_content(file_bytes)
    
    # Check file extension first (more reliable than MIME type)
    if file_name_lower.endswith(".pdf") or file_type_lower.endswith("pdf") or detected_type == "pdf":
        return extract_text_from_pdf(file_bytes)
    elif file_name_lower.endswith(".doc") or detected_type == "doc":
        # Old Word format (.doc) - use LibreOffice
        return extract_text_from_doc(file_bytes)
    elif file_name_lower.endswith(".docx") or file_type_lower == "application/vnd.openxmlformats-officedocument.wordprocessingml.document" or detected_type == "docx":
        # New Word format (.docx) - use python-docx
        # But first check if it's actually a valid DOCX (zip file)
        try:
            import zipfile
            zipfile.ZipFile(BytesIO(file_bytes))
            return extract_text_from_docx(file_bytes)
        except Exception as e:
            # Not a valid DOCX, might be a .doc file misidentified or plain text
            logger.warning(f"File {file_name} identified as DOCX but is not a valid zip file. Detected type: {detected_type}")
            
            # If detected as text, try to read as text
            if detected_type == "text":
                try:
                    return file_bytes.decode('utf-8')
                except:
                    pass
            
            # Try as DOC file
            logger.info(f"Trying to process {file_name} as DOC file")
            return extract_text_from_doc(file_bytes)
    elif file_type_lower == "application/msword":
        # MIME type says .doc
        return extract_text_from_doc(file_bytes)
    elif file_name_lower.endswith(("xlsx", "xls")) or file_type_lower in ["application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             "application/vnd.ms-excel"] or detected_type == "xlsx":
        return extract_text_from_excel(file_bytes)
    elif file_type_lower.startswith("text/") or detected_type == "text" or file_name_lower.endswith(".txt"):
        # Plain text files
        try:
            return file_bytes.decode('utf-8')
        except UnicodeDecodeError:
            try:
                return file_bytes.decode('latin-1')
            except Exception as e:
                return f"Error decoding text file: {str(e)}"
    else:
        # Last resort: try to decode as text if it looks like text
        if detected_type == "text":
            try:
                return file_bytes.decode('utf-8')
            except:
                pass
        return f"Unsupported file type: {file_type}. Detected type: {detected_type}. Supported types: PDF, DOC, DOCX, TXT, XLSX"

def process_uploaded_file(uploaded_file) -> Optional[Dict]:
    """Process an uploaded file and return document info with memory limits"""
    try:
        # Limit file size to prevent memory issues (50MB max)
        max_file_size = 50 * 1024 * 1024  # 50MB
        file_bytes = uploaded_file.read()
        
        if len(file_bytes) > max_file_size:
            error_msg = f"File {uploaded_file.name} is too large ({len(file_bytes) / 1024 / 1024:.1f}MB). Maximum size is 50MB."
            logger.error(error_msg)
            st.error(error_msg)
            return None
        
        file_type = uploaded_file.type
        file_name = uploaded_file.name
        
        # Extract text from file (pass file_name for better detection)
        text_content = extract_text_from_file(file_bytes, file_type, file_name)
        
        # Limit extracted text content to prevent memory issues (100KB max per document)
        max_text_length = 100 * 1024  # 100KB
        if len(text_content) > max_text_length:
            text_content = text_content[:max_text_length] + f"\n\n[Document content truncated - extracted {max_text_length} characters of {len(text_content)} total]"
            logger.warning(f"Document {file_name} content truncated from {len(text_content)} to {max_text_length} characters")
        
        return {
            "name": file_name,
            "type": file_type,
            "size": len(file_bytes),
            "content": text_content,
            "preview": text_content[:500] + "..." if len(text_content) > 500 else text_content
        }
    except MemoryError:
        error_msg = f"Not enough memory to process file {uploaded_file.name}. Please try a smaller file."
        logger.error(error_msg)
        st.error(error_msg)
        return None
    except Exception as e:
        logger.error(f"Error processing uploaded file: {str(e)}")
        st.error(f"Error processing file: {str(e)}")
        return None

# Create a simple synchronous wrapper for async functions
def run_async(coro):
    try:
        return asyncio.run(coro)
    except RuntimeError as e:
        if "This event loop is already running" in str(e):
            # If the event loop is already running, use the current one
            loop = asyncio.get_event_loop()
            return loop.run_until_complete(coro)
        else:
            raise

def initialize_agent():
    """No longer needed - using Flask API instead"""
    st.session_state.agent_running = True
    logger.info("Using Flask API for all operations")
    return True

def end_agent_session():
    """Stop any running deployed services"""
    try:
        # Stop any running deployed services
        if st.session_state.deployer_agent:
            run_async(st.session_state.deployer_agent.stop())
            st.session_state.deployer_agent = None
            st.session_state.backend_url = None
            st.session_state.frontend_url = None
            logger.info("Deployed services stopped")
            return True
    except Exception as e:
        logger.error(f"Error stopping services: {str(e)}")
        return False
    return True

async def call_api_endpoint(endpoint: str, payload: dict):
    """Call a Flask API endpoint asynchronously with memory-efficient handling"""
    # Limit payload size to prevent memory issues
    if "message" in payload:
        max_message_length = 15000  # Limit total message length
        if len(payload["message"]) > max_message_length:
            logger.warning(f"Message too long ({len(payload['message'])} chars), truncating to {max_message_length}")
            payload["message"] = payload["message"][:max_message_length] + "\n\n[Message truncated due to size limits]"
    
    async with httpx.AsyncClient(timeout=600.0) as client:  # Increased timeout to 10 minutes
        try:
            response = await client.post(f"{API_BASE_URL}{endpoint}", json=payload)
            response.raise_for_status()
            return response.json()
        except httpx.HTTPStatusError as e:
            # Try to get error details from response
            error_detail = "Unknown error"
            try:
                error_response = e.response.json()
                error_detail = error_response.get("detail", str(e))
                logger.error(f"API returned error: {error_detail}")
            except:
                error_detail = f"HTTP {e.response.status_code}: {e.response.text[:500]}"
                logger.error(f"API error (non-JSON): {error_detail}")
            raise Exception(f"API Error: {error_detail}")
        except httpx.HTTPError as e:
            logger.error(f"HTTP error calling {endpoint}: {str(e)}")
            raise Exception(f"Connection error: {str(e)}")
        except Exception as e:
            logger.error(f"Error calling {endpoint}: {str(e)}")
            raise

async def get_requirements_analysis(message):
    """Get requirements analysis via Flask API"""
    try:
        result = await call_api_endpoint("/api/analyze-requirements", {
            "message": message,
            "output_format": "text"
        })
        return result.get("result", "")
    except Exception as e:
        logger.error(f"Error getting requirements analysis: {str(e)}")
        return f"Error analyzing requirements: {str(e)}"

async def direct_requirements_to_code(message):
    """
    Complete chatbot creation workflow:
    1. Analyze requirements from user input (and documents if provided)
    2. Generate backend code (Flask/Python)
    3. Generate UI code (React/TailwindCSS) 
    4. Integrate into a complete project structure
    5. Deploy to local servers
    
    This enables a chatbot to create another chatbot seamlessly.
    """
    logger.info(f"Analyzing requirements and generating code via Flask API for: {message[:50]}...")
    
    try:
        # Use Flask API full workflow endpoint - this orchestrates all agents
        logger.info("Calling Flask API /api/generate-full-project endpoint")
        result = await call_api_endpoint("/api/generate-full-project", {
            "message": message,
            "output_format": "text"
        })
        
        if result.get("status") != "success":
            error_msg = result.get("detail", "Unknown error")
            logger.error(f"Flask API workflow failed: {error_msg}")
            return f"Error generating project: {error_msg}", None
        
        # Extract results
        req_analysis = result.get("requirements_analysis", {})
        req_text = req_analysis.get("text", "Requirements analysis completed")
        req_json = req_analysis.get("json", {})
        
        generated = result.get("generated_code", {})
        backend_code = generated.get("backend", "")
        ui_code = generated.get("ui", "")
        
        project_info = result.get("project", {})
        project_dir = project_info.get("directory")
        
        deployment = result.get("deployment", {})
        backend_url = deployment.get("backend_url")
        frontend_url = deployment.get("frontend_url")
        
        # Store deployment URLs if available
        if backend_url and frontend_url:
            st.session_state.backend_url = backend_url
            st.session_state.frontend_url = frontend_url
            logger.info(f"Deployment successful - Backend: {backend_url}, Frontend: {frontend_url}")
        
        # Format project info
        project_info_text = ""
        if project_dir and project_info.get("exists"):
            deployment_info = ""
            if st.session_state.deploy_services and backend_url and frontend_url:
                deployment_info = f"""
## Deployment
Your application has been deployed and is running at:

- Backend API: [{backend_url}]({backend_url})
- Frontend UI: [{frontend_url}]({frontend_url})

The services will remain running until you close this application or click "Stop Services" in the sidebar.
"""
            
            project_info_text = f"""
## Project Integration
A complete project has been assembled at: `{project_dir}`

- Backend code is in the `backend/` directory
- Frontend code is in the `frontend/` directory
- A README.md with setup instructions is included
{deployment_info}
"""
        
        # Return results
        if ui_code and len(ui_code.strip()) > 10:
            return req_text, {
                "backend_code": backend_code,
                "ui_code": ui_code,
                "project_info": project_info_text
            }
        else:
            return req_text, backend_code
            
    except Exception as e:
        logger.error(f"Error in Flask API workflow: {str(e)}")
        return f"I couldn't properly generate your project due to an error: {str(e)}", None

def _check_if_ui_needed(requirements_json, requirements_text):
    """Check if UI generation is needed based on requirements"""
    # Check if requirements explicitly mention UI
    ui_keywords = ["UI", "interface", "frontend", "react", "vue", "angular", 
                  "web page", "website", "responsive", "user interface", 
                  "dashboard", "display", "visualization"]
    
    # Check in keys of the requirements JSON
    if any(key.lower() in ["ui", "ui_components", "design", "design_preferences", 
                          "interface", "frontend", "display"] 
          for key in requirements_json.keys()):
        return True
        
    # Check in JSON values (flattened)
    flat_values = []
    for values in requirements_json.values():
        if isinstance(values, list):
            flat_values.extend([str(v).lower() for v in values])
        else:
            flat_values.append(str(values).lower())
    
    if any(keyword.lower() in " ".join(flat_values) for keyword in ui_keywords):
        return True
        
    # Check in full text
    if requirements_text and any(keyword.lower() in requirements_text.lower() for keyword in ui_keywords):
        return True
        
    return False

def get_agent_response(message, is_code_generation=False):
    """Get a response via Flask API (synchronous wrapper)"""
    try:
        # Call Flask API for requirements analysis
        result = run_async(call_api_endpoint("/api/analyze-requirements", {
            "message": message,
            "output_format": "text"
        }))
        analysis = result.get("result", "")
        
        # For now, return the analysis as the response
        # You can enhance this later with a dedicated chat endpoint if needed
        if analysis:
            return f"Based on your request: {message}\n\n{analysis}"
        else:
            return f"I received your message: {message}\n\nPlease provide more details about what you'd like me to help with."
    except Exception as e:
        logger.error(f"Error getting response from Flask API: {str(e)}")
        return f"Error communicating with Flask API: {str(e)}\n\nPlease ensure Flask API is running at {API_BASE_URL}"

# Main application header
st.title("🤖 Mother of Bots - Multi-Agent Chat Interface")
st.subheader(f"Using {GEMINI_MODEL} via Vertex AI 🦜️")

# Sidebar with info and controls
with st.sidebar:
    st.markdown("## About")
    st.markdown("""
    This is a chat interface for the Mother of Bots multi-agent system. 
    It uses **LangChain** with Google Vertex AI (Gemini) for language model responses and **Flask** for agent orchestration.
    
    **Powered by LangChain** 🦜️ **Vertex AI** ☁️ **Flask** 🌶️
    """)
    
    st.markdown("## Flask API Status")
    api_status = st.empty()
    try:
        async def check_api():
            async with httpx.AsyncClient(timeout=5.0) as client:
                response = await client.get(f"{API_BASE_URL}/health")
                return response.status_code == 200
        api_healthy = run_async(check_api())
        if api_healthy:
            api_status.success(f"✅ Flask API is running at {API_BASE_URL}")
        else:
            api_status.error(f"❌ Flask API not responding at {API_BASE_URL}")
    except Exception as e:
        api_status.error(f"❌ Cannot connect to Flask API: {str(e)}")
        st.warning(f"Please ensure Flask API is running:\n`python -m mother_of_bots.api`")
    
    st.markdown("## Interface Settings")
    show_analysis = st.checkbox("Show requirements analysis", value=st.session_state.show_analysis, 
                               help="Display the requirements analysis step in the conversation")
    
    auto_code = st.checkbox("Auto-generate code", value=st.session_state.auto_generate_code,
                           help="Automatically generate code when requirements suggest a code generation task")
    
    deploy_services = st.checkbox("Deploy generated projects", value=st.session_state.deploy_services,
                                 help="Automatically deploy generated projects to local servers (enabled by default)")
    
    if show_analysis != st.session_state.show_analysis:
        st.session_state.show_analysis = show_analysis
    
    if auto_code != st.session_state.auto_generate_code:
        st.session_state.auto_generate_code = auto_code
    
    if deploy_services != st.session_state.deploy_services:
        st.session_state.deploy_services = deploy_services
    
    # Add a section to show deployed services if available
    if st.session_state.backend_url and st.session_state.frontend_url:
        st.markdown("## Deployed Services")
        st.success("Your application is running!")
        st.markdown(f"- [Backend API]({st.session_state.backend_url})")
        st.markdown(f"- [Frontend UI]({st.session_state.frontend_url})")
        
        if st.button("Stop Services"):
            if st.session_state.deployer_agent:
                run_async(st.session_state.deployer_agent.stop())
                st.session_state.deployer_agent = None
                st.session_state.backend_url = None
                st.session_state.frontend_url = None
                st.success("Services stopped successfully")
                st.rerun()
    
    st.markdown("## LangChain + Vertex AI Status")
    langchain_status = st.empty()
    
    # Check LangChain Vertex AI availability
    try:
        from langchain_google_vertexai import ChatVertexAI
        
        # Try to initialize LangChain Vertex AI to verify setup
        try:
            test_llm = ChatVertexAI(
                model=GEMINI_MODEL,
                project=GCP_PROJECT_ID,
                location=GCP_LOCATION
            )
            # If initialization succeeds, LangChain + Vertex AI is ready
            langchain_status.success("✅ LangChain + Vertex AI is active and ready")
            st.info(f"**Using LangChain** with Vertex AI ({GEMINI_MODEL})\n\nProject: {GCP_PROJECT_ID}\nLocation: {GCP_LOCATION}\n\nAll LLM operations use LangChain:\n- Requirements Analysis\n- Code Generation\n- UI Generation\n- User Interactions")
        except Exception as e:
            langchain_status.warning(f"⚠️ LangChain initialized but Vertex AI connection test failed: {str(e)}")
            st.warning(f"LangChain is installed but may have issues with Vertex AI:\n`{str(e)}`\n\nPlease verify:\n- GCP credentials are set up (run `gcloud auth application-default login`)\n- Vertex AI API is enabled in project '{GCP_PROJECT_ID}'\n- Model '{GEMINI_MODEL}' is available")
    except ImportError as e:
        langchain_status.error(f"❌ LangChain Vertex AI not available: {str(e)}")
        st.warning("Please install LangChain Vertex AI: `pip install langchain-google-vertexai`")
    
    st.markdown("## GCP / Vertex AI Status")
    vertex_status = st.empty()
    
    # Check GCP credentials
    try:
        import google.auth
        credentials, project = google.auth.default()
        if credentials:
            vertex_status.success(f"✅ GCP credentials configured (Project: {project or GCP_PROJECT_ID})")
        else:
            vertex_status.warning("⚠️ GCP credentials not found")
            st.warning("Run `gcloud auth application-default login` to authenticate")
    except Exception as e:
        vertex_status.error(f"❌ GCP auth check failed: {str(e)}")
        st.warning("Please ensure GCP SDK is installed and run:\n`gcloud auth application-default login`")
    
    
    st.markdown("## Settings")
    if st.button("Reset Conversation"):
        st.session_state.messages = []
        st.rerun()
    
    # Code Generation info
    st.sidebar.header("Code Generation")
    st.sidebar.write("""
    This system automatically detects what code needs to be generated without requiring you to specify.
    
    The process uses three specialized components:
    
    1. **Backend Code Generator**: Produces Python code for server logic, APIs, and database models
    
    2. **UI Generator**: Creates React/TailwindCSS code for frontend components
    
    3. **Project Integrator**: Combines backend and frontend into a structured project:
      - Backend directory with Python code and dependencies
      - Frontend directory with React components
      - Configuration for connecting frontend to backend API
      - README with setup instructions
    
    4. **Project Deployer**: Handles deployment of the integrated project:
      - Starts the backend API server using Uvicorn
      - Serves the frontend UI using a simple HTTP server
      - Provides live URLs to access both services
      
    Deployment is enabled by default - your application will automatically run on localhost when generated.
    """)
    
    st.sidebar.info("Flask API mode is processing your request.")
    
    st.markdown("## System Status")
    if st.session_state.agent_running:
        st.success("✅ System is ready (Flask API mode)")
    else:
        st.warning("System not initialized")
        if st.button("Initialize System"):
            initialize_agent()
            st.rerun()

# Initialize agent on page load
if not st.session_state.agent_running:
    try:
        initialize_agent()
    except Exception as e:
        st.error(f"Failed to initialize agent: {str(e)}")
        logger.error(f"Agent initialization error: {str(e)}")

# Display chat messages
for i, message in enumerate(st.session_state.messages):
    if message["role"] == "user":
        avatar = "🧑‍💻"
    elif message["role"] == "assistant":
        avatar = "🤖"
    else:  # system message for requirements analysis
        avatar = "🔎"
        
    with st.container():
        # Display user messages with document attachments
        if message["role"] == "user" and message.get("documents"):
            doc_badges = " ".join([f"📎 {doc}" for doc in message.get("documents", [])])
            message_content = f"{message['content']}\n\n<div style='margin-top: 0.5rem; font-size: 0.85em; opacity: 0.8;'>{doc_badges}</div>"
            st.markdown(f"""
            <div class="chat-message {message['role']}">
                <div class="message">
                    <b>{avatar} {message['role'].title()}</b>
                    <br>
                    {message_content}
                </div>
            </div>
            """, unsafe_allow_html=True)
            continue
        
        # Special handling for requirements analysis (system messages)
        if message["role"] == "system" and "Requirements Analysis" in message["content"]:
            st.markdown(f"""
            <div class="chat-message system">
                <div class="message requirements-analysis">
                    <b>{avatar} Requirements Analysis</b>
                    <br>
                    {message["content"].replace("**Requirements Analysis:**", "", 1).strip()}
                </div>
            </div>
            """, unsafe_allow_html=True)
        # Special handling for code blocks in assistant messages
        elif message["role"] == "assistant" and "```" in message["content"]:
            # Format code blocks for better display
            content = message["content"]
            
            # Check if this is a code generation result with markdown headers
            if "## Requirements Analysis" in content and ("## Generated Code" in content or "## Generated Backend Code" in content):
                # This is a code generation result, use special formatting
                parts = content.split("## ")
                formatted_content = ""
                
                for part in parts:
                    if part.strip():
                        if part.startswith("Requirements Analysis"):
                            # Format requirements section
                            section_title = "Requirements Analysis"
                            section_content = part.replace("Requirements Analysis", "", 1).strip()
                            formatted_content += f'<h2>{section_title}</h2>\n{section_content}\n'
                        elif part.startswith("Generated Code"):
                            # Format code section
                            section_title = "Generated Code"
                            section_content = part.replace("Generated Code", "", 1).strip()
                            formatted_content += f'<h2>{section_title}</h2>\n{section_content}\n'
                        elif part.startswith("Generated Backend Code"):
                            # Format backend code section
                            section_title = "Generated Backend Code"
                            section_content = part.replace("Generated Backend Code", "", 1).strip()
                            formatted_content += f'<h2>{section_title}</h2>\n{section_content}\n'
                        elif part.startswith("Generated UI Code"):
                            # Format UI code section
                            section_title = "Generated UI Code"
                            section_content = part.replace("Generated UI Code", "", 1).strip()
                            formatted_content += f'<h2>{section_title}</h2>\n{section_content}\n'
                        else:
                            # Regular content
                            formatted_content += part
                
                st.markdown(f"""
                <div class="chat-message assistant">
                    <div class="message code-generation-output">
                        <b>{avatar} {message["role"].title()}</b>
                        <br>
                        {formatted_content}
                    </div>
                </div>
                """, unsafe_allow_html=True)
            else:
                # Regular message with code blocks, just display normally
                st.markdown(f"""
                <div class="chat-message {message["role"]}">
                    <div class="message">
                        <b>{avatar} {message["role"].title()}</b>
                        <br>
                        {content}
                    </div>
                </div>
                """, unsafe_allow_html=True)
        else:
            # Regular message display
            st.markdown(f"""
            <div class="chat-message {message["role"]}">
                <div class="message">
                    <b>{avatar} {message["role"].title()}</b>
                    <br>
                    {message["content"]}
                </div>
            </div>
            """, unsafe_allow_html=True)

# File uploader section
st.markdown("### 📎 Upload Documents")
uploaded_files = st.file_uploader(
    "Upload documents (PDF, DOCX, TXT, XLSX)",
    type=["pdf", "docx", "doc", "txt", "xlsx", "xls"],
    accept_multiple_files=True,
    help="Upload documents to include in the conversation context"
)

# Process uploaded files
if uploaded_files:
    for uploaded_file in uploaded_files:
        # Check if file is already processed
        file_already_uploaded = any(doc["name"] == uploaded_file.name for doc in st.session_state.uploaded_documents)
        
        if not file_already_uploaded:
            with st.spinner(f"Processing {uploaded_file.name}..."):
                doc_info = process_uploaded_file(uploaded_file)
                if doc_info:
                    st.session_state.uploaded_documents.append(doc_info)
                    st.success(f"✅ {uploaded_file.name} processed successfully ({len(doc_info['content'])} characters)")

# Display uploaded documents
if st.session_state.uploaded_documents:
    with st.expander(f"📚 Uploaded Documents ({len(st.session_state.uploaded_documents)})", expanded=False):
        for i, doc in enumerate(st.session_state.uploaded_documents):
            col1, col2 = st.columns([3, 1])
            with col1:
                st.markdown(f"**{doc['name']}** ({doc['type']}, {doc['size']} bytes)")
                st.text_area(
                    "Preview",
                    doc['preview'],
                    height=100,
                    key=f"doc_preview_{i}",
                    disabled=True
                )
            with col2:
                if st.button("Remove", key=f"remove_doc_{i}"):
                    st.session_state.uploaded_documents.pop(i)
                    st.rerun()

# Chat input
user_input = st.chat_input("Type your message here...")

# Process user input
if user_input and not st.session_state.waiting_for_response:
    # Check total message size before processing
    estimated_size = len(user_input)
    if st.session_state.uploaded_documents:
        for doc in st.session_state.uploaded_documents:
            estimated_size += len(doc.get('content', ''))
    
    # Warn if message is very large
    if estimated_size > 50000:  # 50KB
        st.warning(f"⚠️ Large message detected ({estimated_size:,} characters). Processing may take longer and use more memory.")
    
    # Prepare message with document context
    message_content = user_input
    attached_docs = []
    
    # Include document content if documents are uploaded
    # Limit document content to prevent memory issues and message length problems
    if st.session_state.uploaded_documents:
        doc_context = "\n\n--- Uploaded Documents Context ---\n"
        max_doc_length = 3000  # Reduced to 3000 chars per document to prevent memory issues
        max_total_length = 10000  # Maximum total document context length
        total_length = 0
        
        for doc in st.session_state.uploaded_documents:
            if total_length >= max_total_length:
                doc_context += f"\n[Document: {doc['name']} - Content omitted due to size limits]\n"
                continue
                
            doc_content = doc['content']
            remaining_space = max_total_length - total_length
            doc_max_length = min(max_doc_length, remaining_space - 200)  # Reserve space for formatting
            
            if len(doc_content) > doc_max_length:
                doc_content = doc_content[:doc_max_length] + f"\n\n[Document truncated - showing first {doc_max_length} characters of {len(doc['content'])} total]"
            
            doc_context += f"\n[Document: {doc['name']}]\n{doc_content}\n"
            total_length += len(doc_content)
            
            if total_length >= max_total_length:
                doc_context += "\n[Additional documents omitted due to size limits]"
                break
        
        message_content = f"{user_input}\n{doc_context}"
        attached_docs = [doc['name'] for doc in st.session_state.uploaded_documents]
    
    # Add user message to chat
    message_data = {
        "role": "user",
        "content": user_input,
        "documents": attached_docs
    }
    st.session_state.messages.append(message_data)
    
    # Set waiting flag
    st.session_state.waiting_for_response = True
    
    # Rerun to display user message immediately
    st.rerun()

# Process response (after rerun)
if st.session_state.waiting_for_response:
    # Force garbage collection before processing to free memory
    gc.collect()
    
    with st.status("Processing...", expanded=True) as status:
        # Get the last user message with document context
        last_user_message_obj = next((msg for msg in reversed(st.session_state.messages) 
                                     if msg["role"] == "user"), None)
        
        if last_user_message_obj:
            last_user_message = last_user_message_obj["content"]
            # Include document content if documents were attached
            # Limit document content to prevent memory issues and message length problems
            if last_user_message_obj.get("documents") and st.session_state.uploaded_documents:
                doc_context = "\n\n--- Uploaded Documents Context ---\n"
                max_doc_length = 3000  # Reduced to 3000 chars per document
                max_total_length = 10000  # Maximum total document context length
                total_length = 0
                
                for doc in st.session_state.uploaded_documents:
                    if doc['name'] in last_user_message_obj.get("documents", []):
                        if total_length >= max_total_length:
                            doc_context += f"\n[Document: {doc['name']} - Content omitted due to size limits]\n"
                            continue
                            
                        doc_content = doc['content']
                        remaining_space = max_total_length - total_length
                        doc_max_length = min(max_doc_length, remaining_space - 200)  # Reserve space for formatting
                        
                        if len(doc_content) > doc_max_length:
                            doc_content = doc_content[:doc_max_length] + f"\n\n[Document truncated - showing first {doc_max_length} characters of {len(doc['content'])} total]"
                        
                        doc_context += f"\n[Document: {doc['name']}]\n{doc_content}\n"
                        total_length += len(doc_content)
                        
                        if total_length >= max_total_length:
                            doc_context += "\n[Additional documents omitted due to size limits]"
                            break
                
                last_user_message = f"{last_user_message}\n{doc_context}"
        else:
            last_user_message = ""
        
        # Determine if this is a code generation request
        code_keywords = ["generate code", "create code", "write code", "code for", "generate a program", 
                         "build an application", "develop a system", "create an app", "write a program",
                         "script for", "implement a solution", "code that can", "build a website",
                         "create a function", "make an algorithm", "create a chatbot", "create chatbot",
                         "build a chatbot", "build chatbot", "create an application", "create application",
                         "based on the attached document", "based on the document", "implement"]
        
        # First check if there are explicit code generation keywords
        is_code_request = any(keyword in last_user_message.lower() for keyword in code_keywords)
        
        # If not explicitly a code request, do a more thorough analysis
        if not is_code_request and st.session_state.auto_generate_code:
            try:
                # Get full requirements analysis
                req_analysis = run_async(get_requirements_analysis(last_user_message))
                
                # Look for code-related keywords in the requirements analysis
                code_indicator_phrases = ["code", "program", "application", "function", "module", "class", 
                                         "API", "endpoint", "system", "backend", "frontend", "algorithm",
                                         "software", "app", "website", "interface", "database"]
                
                is_likely_code_request = any(indicator in req_analysis.lower() for indicator in code_indicator_phrases)
                
                # Count the number of indicators found to determine confidence
                indicator_count = sum(1 for indicator in code_indicator_phrases if indicator in req_analysis.lower())
                
                # If at least 2 code indicators are found, treat as a code request
                if is_likely_code_request and indicator_count >= 2:
                    logger.info(f"Requirements analysis suggests this is a code-related request (found {indicator_count} indicators)")
                    is_code_request = True
                    
                # Also check for specific requirement categories that suggest code generation
                if "functionalities" in req_analysis.lower() and any(tech in req_analysis.lower() 
                                                                 for tech in ["python", "javascript", "java", "api", "database"]):
                    logger.info("Requirements mention technical functionalities, treating as code request")
                    is_code_request = True
            except Exception as e:
                logger.error(f"Error in code requirements detection: {str(e)}")
        
        try:
            if is_code_request and st.session_state.auto_generate_code:
                # Direct code generation path
                logger.info("Detected code generation request, processing directly")
                
                status.update(label="Step 1/5: Analyzing requirements...", state="running")
                st.write("🤖 Creating your chatbot...")
                st.write("📋 Step 1: Analyzing requirements")
                
                try:
                    # Directly analyze requirements and generate code without intermediate steps
                    status.update(label="Step 2/5: Generating backend code...", state="running")
                    st.write("⚙️ Step 2: Generating backend code")
                    
                    status.update(label="Step 3/5: Generating UI...", state="running")
                    st.write("🎨 Step 3: Generating user interface")
                    
                    status.update(label="Step 4/5: Integrating project...", state="running")
                    st.write("🔗 Step 4: Integrating components")
                    
                    status.update(label="Step 5/5: Deploying...", state="running")
                    st.write("🚀 Step 5: Deploying your chatbot")
                    
                    requirements_text, generated_code = run_async(direct_requirements_to_code(last_user_message))
                    
                    if generated_code:
                        # Check if we received a dict with both backend and UI code
                        if isinstance(generated_code, dict) and "backend_code" in generated_code and "ui_code" in generated_code:
                            # Format the response with both requirements, backend code, and UI code
                            response = f"""## Requirements Analysis
{requirements_text}

## Generated Backend Code (Python)
```python
{generated_code['backend_code']}
```

## Generated Frontend UI (React)
```jsx
{generated_code['ui_code']}
```

{generated_code.get('project_info', '')}
"""
                        else:
                            # Format the response with both requirements and code (backend only)
                            response = f"""## Requirements Analysis
{requirements_text}

## Generated Backend Code (Python)
```python
{generated_code}
```
"""
                        
                        logger.info("Flask API code generation completed successfully")
                    else:
                        # Fallback to normal response if code generation failed
                        logger.warning("Code generation failed, falling back to normal response")
                        response = get_agent_response(last_user_message)
                except Exception as e:
                    error_msg = str(e)
                    logger.error(f"Error in code generation workflow: {error_msg}")
                    
                    # Provide helpful error message
                    if "API Error" in error_msg or "Connection error" in error_msg:
                        response = f"""## ⚠️ Error Generating Chatbot

I encountered an error while creating your chatbot:

**Error:** {error_msg}

**Troubleshooting:**
1. Make sure the Flask API server is running: `python -m mother_of_bots.api`
2. Check that GCP credentials are set up: `gcloud auth application-default login`
3. Verify Vertex AI API is enabled in project '{GCP_PROJECT_ID}'
4. Try simplifying your request or breaking it into smaller parts
5. Check the API logs for more details

**Your request was:** {last_user_message[:200]}...
"""
                    else:
                        response = f"""## ⚠️ Error Generating Chatbot

I encountered an error: {error_msg}

Please try again or simplify your request.
"""
            else:
                # Regular chat path with separate requirements analysis
                if st.session_state.show_analysis:
                    st.write("Step 1: Analyzing requirements using LangChain...")
                    status.update(label="Analyzing requirements with LangChain...", state="running")
                    
                    # Get requirements analysis
                    requirements_analysis = run_async(get_requirements_analysis(last_user_message))
                    
                    # Add system message for requirements analysis
                    st.session_state.messages.append({
                        "role": "system", 
                        "content": f"**Requirements Analysis:**\n\n{requirements_analysis}"
                    })
                    
                    # Update status for a temporary pause to show analysis
                    status.update(label="Analysis complete", state="complete")
                    time.sleep(0.5)  # Small pause to let user see the analysis
                    st.rerun()  # Rerun to show the analysis before generating response
                
                # Generate regular response
                st.write("Generating response using LangChain...")
                status.update(label="Generating response with LangChain...", state="running")
                response = get_agent_response(last_user_message)
            
            # Add bot message to chat
            st.session_state.messages.append({"role": "assistant", "content": response})
            
            if is_code_request and st.session_state.auto_generate_code:
                status.update(label="Code generation complete!", state="complete", expanded=False)
            else:
                status.update(label="Response ready!", state="complete", expanded=False)
        except Exception as e:
            st.error(f"Error generating response: {str(e)}")
            logger.error(f"Response generation error: {str(e)}")
            st.session_state.messages.append({"role": "assistant", "content": f"I'm sorry, I encountered an error: {str(e)}"})
        finally:
            # Reset waiting flag
            st.session_state.waiting_for_response = False
            # Force garbage collection after processing
            gc.collect()
    
    # Rerun to display bot message
    st.rerun()

# Register a cleanup function
def cleanup():
    if st.session_state.agent_running:
        end_agent_session()
    
    # Also stop any running deployed services
    if st.session_state.deployer_agent:
        run_async(st.session_state.deployer_agent.stop())
        st.session_state.deployer_agent = None

# Register the cleanup with streamlit
import atexit
atexit.register(cleanup) 